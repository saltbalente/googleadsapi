"""
📦 EXPORT MANAGER - Gestor de Exportación Avanzada
Sistema completo de exportación de anuncios a múltiples formatos
Versión: 2.0
Fecha: 2025-01-13
Autor: saltbalente
"""

import logging
import json
import csv
from typing import Dict, List, Optional, Any, Union
from datetime import datetime
from pathlib import Path
from io import StringIO, BytesIO
import zipfile

# Configurar logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class ExportManager:
    """
    Gestor de exportación que proporciona:
    - Exportación a múltiples formatos (CSV, JSON, Excel, XML, TXT)
    - Exportación para Google Ads Editor
    - Exportación para Facebook Ads
    - Exportación bulk de múltiples anuncios
    - Templates personalizables
    - Compresión de archivos
    - Validación pre-exportación
    - Historial de exportaciones
    - Exportación programada
    """
    
    # =========================================================================
    # CONFIGURACIÓN DE FORMATOS
    # =========================================================================
    
    SUPPORTED_FORMATS = [
        'csv',           # CSV estándar
        'json',          # JSON
        'excel',         # Excel (.xlsx) - requiere openpyxl
        'xml',           # XML
        'txt',           # Texto plano
        'google_ads',    # CSV para Google Ads Editor
        'facebook_ads',  # CSV para Facebook Ads
        'zip'           # ZIP con múltiples formatos
    ]
    
    # Templates de columnas para diferentes formatos
    GOOGLE_ADS_COLUMNS = [
        'Campaign',
        'Ad Group',
        'Headline 1',
        'Headline 2',
        'Headline 3',
        'Headline 4',
        'Headline 5',
        'Headline 6',
        'Headline 7',
        'Headline 8',
        'Headline 9',
        'Headline 10',
        'Headline 11',
        'Headline 12',
        'Headline 13',
        'Headline 14',
        'Headline 15',
        'Description 1',
        'Description 2',
        'Description 3',
        'Description 4',
        'Path 1',
        'Path 2',
        'Final URL',
        'Status'
    ]
    
    FACEBOOK_ADS_COLUMNS = [
        'Campaign Name',
        'Ad Set Name',
        'Ad Name',
        'Primary Text',
        'Headline',
        'Description',
        'Destination URL',
        'Call to Action',
        'Status'
    ]
    
    def __init__(
        self,
        export_dir: Optional[str] = None,
        auto_create_dirs: bool = True,
        include_timestamp: bool = True,
        default_format: str = 'csv'
    ):
        """
        Inicializa el gestor de exportación.
        
        Args:
            export_dir: Directorio base para exportaciones
            auto_create_dirs: Crear directorios automáticamente
            include_timestamp: Incluir timestamp en nombres de archivo
            default_format: Formato por defecto
        """
        # Configurar directorio de exportación
        if export_dir:
            self.export_dir = Path(export_dir)
        else:
            self.export_dir = Path(__file__).parent.parent / "exports"
        
        if auto_create_dirs:
            self.export_dir.mkdir(parents=True, exist_ok=True)
        
        self.include_timestamp = include_timestamp
        self.default_format = default_format
        
        # Historial de exportaciones
        self.export_history: List[Dict[str, Any]] = []
        
        # Estadísticas
        self.stats = {
            'total_exports': 0,
            'exports_by_format': {},
            'total_files_created': 0,
            'total_bytes_exported': 0
        }
        
        logger.info(f"✅ ExportManager inicializado")
        logger.info(f"   - Export dir: {self.export_dir}")
        logger.info(f"   - Default format: {default_format}")
        logger.info(f"   - Supported formats: {', '.join(self.SUPPORTED_FORMATS)}")
    
    # =========================================================================
    # EXPORTACIÓN PRINCIPAL
    # =========================================================================
    
    def export(
        self,
        ads: Union[Dict[str, Any], List[Dict[str, Any]]],
        format: str = 'csv',
        filename: Optional[str] = None,
        **options
    ) -> Dict[str, Any]:
        """
        Exporta anuncios al formato especificado.
        
        Args:
            ads: Anuncio o lista de anuncios
            format: Formato de exportación
            filename: Nombre del archivo (sin extensión)
            **options: Opciones adicionales según formato
        
        Returns:
            Diccionario con información de la exportación
        """
        export_id = f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        
        logger.info(f"📦 Iniciando exportación: {export_id}")
        logger.info(f"   - Formato: {format}")
        
        # Validar formato
        if format not in self.SUPPORTED_FORMATS:
            return {
                'success': False,
                'error': f"Formato '{format}' no soportado",
                'supported_formats': self.SUPPORTED_FORMATS
            }
        
        # Normalizar entrada
        if isinstance(ads, dict):
            ads_list = [ads]
        else:
            ads_list = ads
        
        if not ads_list:
            return {
                'success': False,
                'error': 'No hay anuncios para exportar'
            }
        
        logger.info(f"   - Anuncios: {len(ads_list)}")
        
        # Validar anuncios
        validation_result = self._validate_ads(ads_list)
        if not validation_result['valid']:
            logger.warning(f"⚠️ Advertencia: {validation_result['warning']}")
        
        # Generar nombre de archivo
        if not filename:
            filename = self._generate_filename(format)
        else:
            filename = self._add_timestamp(filename) if self.include_timestamp else filename
        
        # Exportar según formato
        try:
            if format == 'csv':
                result = self._export_csv(ads_list, filename, **options)
            elif format == 'json':
                result = self._export_json(ads_list, filename, **options)
            elif format == 'excel':
                result = self._export_excel(ads_list, filename, **options)
            elif format == 'xml':
                result = self._export_xml(ads_list, filename, **options)
            elif format == 'txt':
                result = self._export_txt(ads_list, filename, **options)
            elif format == 'google_ads':
                result = self._export_google_ads(ads_list, filename, **options)
            elif format == 'facebook_ads':
                result = self._export_facebook_ads(ads_list, filename, **options)
            elif format == 'zip':
                result = self._export_zip(ads_list, filename, **options)
            else:
                return {
                    'success': False,
                    'error': f"Formato '{format}' no implementado"
                }
            
            # Agregar metadata
            result['export_id'] = export_id
            result['format'] = format
            result['ads_count'] = len(ads_list)
            result['exported_at'] = datetime.now().isoformat()
            result['validation'] = validation_result
            
            # Guardar en historial
            self.export_history.append(result)
            
            # Actualizar estadísticas
            self._update_stats(result)
            
            logger.info(f"✅ Exportación completada: {result['file_path']}")
            
            return result
            
        except Exception as e:
            logger.error(f"❌ Error en exportación: {e}", exc_info=True)
            return {
                'success': False,
                'error': str(e),
                'format': format
            }
    
    # =========================================================================
    # EXPORTACIÓN CSV
    # =========================================================================
    
    def _export_csv(
        self,
        ads: List[Dict[str, Any]],
        filename: str,
        delimiter: str = ',',
        include_header: bool = True,
        encoding: str = 'utf-8-sig'  # BOM para Excel
    ) -> Dict[str, Any]:
        """Exporta a formato CSV."""
        file_path = self.export_dir / f"{filename}.csv"
        
        # Determinar columnas
        all_keys = set()
        for ad in ads:
            # Incluir headlines y descriptions como columnas separadas
            if 'headlines' in ad:
                for i, _ in enumerate(ad['headlines'], 1):
                    all_keys.add(f'headline_{i}')
            
            if 'descriptions' in ad:
                for i, _ in enumerate(ad['descriptions'], 1):
                    all_keys.add(f'description_{i}')
            
            # Otras keys
            for key in ad.keys():
                if key not in ['headlines', 'descriptions']:
                    all_keys.add(key)
        
        # Ordenar columnas
        columns = sorted(all_keys)
        
        # Escribir CSV
        with open(file_path, 'w', newline='', encoding=encoding) as f:
            writer = csv.DictWriter(f, fieldnames=columns, delimiter=delimiter)
            
            if include_header:
                writer.writeheader()
            
            for ad in ads:
                row = {}
                
                # Agregar headlines
                if 'headlines' in ad:
                    for i, headline in enumerate(ad['headlines'], 1):
                        row[f'headline_{i}'] = headline
                
                # Agregar descriptions
                if 'descriptions' in ad:
                    for i, desc in enumerate(ad['descriptions'], 1):
                        row[f'description_{i}'] = desc
                
                # Agregar otras keys
                for key, value in ad.items():
                    if key not in ['headlines', 'descriptions']:
                        # Serializar listas/dicts como JSON
                        if isinstance(value, (list, dict)):
                            row[key] = json.dumps(value, ensure_ascii=False)
                        else:
                            row[key] = value
                
                writer.writerow(row)
        
        file_size = file_path.stat().st_size
        
        return {
            'success': True,
            'file_path': str(file_path),
            'file_name': file_path.name,
            'file_size': file_size,
            'rows': len(ads),
            'columns': len(columns)
        }
    
    # =========================================================================
    # EXPORTACIÓN JSON
    # =========================================================================
    
    def _export_json(
        self,
        ads: List[Dict[str, Any]],
        filename: str,
        indent: int = 2,
        ensure_ascii: bool = False
    ) -> Dict[str, Any]:
        """Exporta a formato JSON."""
        file_path = self.export_dir / f"{filename}.json"
        
        export_data = {
            'exported_at': datetime.now().isoformat(),
            'total_ads': len(ads),
            'ads': ads
        }
        
        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(export_data, f, indent=indent, ensure_ascii=ensure_ascii)
        
        file_size = file_path.stat().st_size
        
        return {
            'success': True,
            'file_path': str(file_path),
            'file_name': file_path.name,
            'file_size': file_size,
            'ads_count': len(ads)
        }
    
    # =========================================================================
    # EXPORTACIÓN EXCEL
    # =========================================================================
    
    def _export_excel(
        self,
        ads: List[Dict[str, Any]],
        filename: str,
        sheet_name: str = 'Anuncios'
    ) -> Dict[str, Any]:
        """Exporta a formato Excel (.xlsx)."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            return {
                'success': False,
                'error': 'openpyxl no instalado. Instala con: pip install openpyxl'
            }
        
        file_path = self.export_dir / f"{filename}.xlsx"
        
        # Crear workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # Determinar columnas
        all_keys = set()
        for ad in ads:
            if 'headlines' in ad:
                for i in range(len(ad['headlines'])):
                    all_keys.add(f'headline_{i+1}')
            if 'descriptions' in ad:
                for i in range(len(ad['descriptions'])):
                    all_keys.add(f'description_{i+1}')
            for key in ad.keys():
                if key not in ['headlines', 'descriptions']:
                    all_keys.add(key)
        
        columns = sorted(all_keys)
        
        # Escribir header
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_num, column in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = column
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Escribir datos
        for row_num, ad in enumerate(ads, 2):
            for col_num, column in enumerate(columns, 1):
                if column.startswith('headline_'):
                    idx = int(column.split('_')[1]) - 1
                    value = ad.get('headlines', [])[idx] if idx < len(ad.get('headlines', [])) else ''
                elif column.startswith('description_'):
                    idx = int(column.split('_')[1]) - 1
                    value = ad.get('descriptions', [])[idx] if idx < len(ad.get('descriptions', [])) else ''
                else:
                    value = ad.get(column, '')
                    if isinstance(value, (list, dict)):
                        value = json.dumps(value, ensure_ascii=False)
                
                ws.cell(row=row_num, column=col_num, value=value)
        
        # Ajustar anchos de columna
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Guardar
        wb.save(file_path)
        
        file_size = file_path.stat().st_size
        
        return {
            'success': True,
            'file_path': str(file_path),
            'file_name': file_path.name,
            'file_size': file_size,
            'rows': len(ads),
            'columns': len(columns),
            'sheet_name': sheet_name
        }
    
    # =========================================================================
    # EXPORTACIÓN XML
    # =========================================================================
    
    def _export_xml(
        self,
        ads: List[Dict[str, Any]],
        filename: str,
        root_tag: str = 'ads',
        item_tag: str = 'ad'
    ) -> Dict[str, Any]:
        """Exporta a formato XML."""
        file_path = self.export_dir / f"{filename}.xml"
        
        xml_lines = ['<?xml version="1.0" encoding="UTF-8"?>']
        xml_lines.append(f'<{root_tag}>')
        xml_lines.append(f'  <exported_at>{datetime.now().isoformat()}</exported_at>')
        xml_lines.append(f'  <total>{len(ads)}</total>')
        xml_lines.append('  <items>')
        
        for ad in ads:
            xml_lines.append(f'    <{item_tag}>')
            
            # Headlines
            if 'headlines' in ad:
                xml_lines.append('      <headlines>')
                for i, headline in enumerate(ad['headlines'], 1):
                    escaped = self._escape_xml(headline)
                    xml_lines.append(f'        <headline id="{i}">{escaped}</headline>')
                xml_lines.append('      </headlines>')
            
            # Descriptions
            if 'descriptions' in ad:
                xml_lines.append('      <descriptions>')
                for i, desc in enumerate(ad['descriptions'], 1):
                    escaped = self._escape_xml(desc)
                    xml_lines.append(f'        <description id="{i}">{escaped}</description>')
                xml_lines.append('      </descriptions>')
            
            # Otras propiedades
            for key, value in ad.items():
                if key not in ['headlines', 'descriptions']:
                    if isinstance(value, (list, dict)):
                        value = json.dumps(value)
                    escaped = self._escape_xml(str(value))
                    xml_lines.append(f'      <{key}>{escaped}</{key}>')
            
            xml_lines.append(f'    </{item_tag}>')
        
        xml_lines.append('  </items>')
        xml_lines.append(f'</{root_tag}>')
        
        # Escribir archivo
        with open(file_path, 'w', encoding='utf-8') as f:
            f.write('\n'.join(xml_lines))
        
        file_size = file_path.stat().st_size
        
        return {
            'success': True,
            'file_path': str(file_path),
            'file_name': file_path.name,
            'file_size': file_size,
            'ads_count': len(ads)
        }
    
    def _escape_xml(self, text: str) -> str:
        """Escapa caracteres especiales para XML."""
        text = str(text)
        text = text.replace('&', '&amp;')
        text = text.replace('<', '&lt;')
        text = text.replace('>', '&gt;')
        text = text.replace('"', '&quot;')
        text = text.replace("'", '&apos;')
        return text
    
    # =========================================================================
    # EXPORTACIÓN TXT
    # =========================================================================
    
    def _export_txt(
        self,
        ads: List[Dict[str, Any]],
        filename: str,
        separator: str = '\n---\n'
    ) -> Dict[str, Any]:
        """Exporta a formato texto plano."""
        file_path = self.export_dir / f"{filename}.txt"
        
        lines = []
        lines.append(f"EXPORTACIÓN DE ANUNCIOS")
        lines.append(f"Fecha: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        lines.append(f"Total de anuncios: {len(ads)}")
        lines.append("=" * 60)
        lines.append("")
        
        for i, ad in enumerate(ads, 1):
            lines.append(f"ANUNCIO #{i}")
            lines.append("-" * 60)
            
            # Headlines
            if 'headlines' in ad:
                lines.append("\nTITULARES:")
                for j, headline in enumerate(ad['headlines'], 1):
                    lines.append(f"  {j}. {headline}")
            
            # Descriptions
            if 'descriptions' in ad:
                lines.append("\nDESCRIPCIONES:")
                for j, desc in enumerate(ad['descriptions'], 1):
                    lines.append(f"  {j}. {desc}")
            
            # Otras propiedades
            other_props = {k: v for k, v in ad.items() 
                          if k not in ['headlines', 'descriptions']}
            
            if other_props:
                lines.append("\nOTRAS PROPIEDADES:")
                for key, value in other_props.items():
                    if isinstance(value, (list, dict)):
                        value = json.dumps(value, indent=2)
                    lines.append(f"  {key}: {value}")
            
            lines.append(separator)
        
        # Escribir archivo
        with open(file_path, 'w', encoding='utf-8') as f:
            f.write('\n'.join(lines))
        
        file_size = file_path.stat().st_size
        
        return {
            'success': True,
            'file_path': str(file_path),
            'file_name': file_path.name,
            'file_size': file_size,
            'ads_count': len(ads)
        }
    
    # =========================================================================
    # EXPORTACIÓN GOOGLE ADS
    # =========================================================================
    
    def _export_google_ads(
        self,
        ads: List[Dict[str, Any]],
        filename: str,
        campaign_name: str = "Nueva Campaña",
        ad_group_name: str = "Grupo de Anuncios 1",
        final_url: str = "https://example.com",
        path1: str = "",
        path2: str = "",
        status: str = "Paused"
    ) -> Dict[str, Any]:
        """
        Exporta en formato compatible con Google Ads Editor.
        
        Args:
            ads: Lista de anuncios
            campaign_name: Nombre de la campaña
            ad_group_name: Nombre del grupo de anuncios
            final_url: URL de destino
            path1: Ruta 1 del display URL
            path2: Ruta 2 del display URL
            status: Estado inicial (Paused/Enabled)
        """
        file_path = self.export_dir / f"{filename}_google_ads.csv"
        
        with open(file_path, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.DictWriter(f, fieldnames=self.GOOGLE_ADS_COLUMNS)
            writer.writeheader()
            
            for ad in ads:
                row = {
                    'Campaign': campaign_name,
                    'Ad Group': ad_group_name,
                    'Path 1': path1,
                    'Path 2': path2,
                    'Final URL': final_url,
                    'Status': status
                }
                
                # Headlines (máximo 15)
                headlines = ad.get('headlines', [])
                for i in range(15):
                    col_name = f'Headline {i+1}'
                    row[col_name] = headlines[i] if i < len(headlines) else ''
                
                # Descriptions (máximo 4)
                descriptions = ad.get('descriptions', [])
                for i in range(4):
                    col_name = f'Description {i+1}'
                    row[col_name] = descriptions[i] if i < len(descriptions) else ''
                
                writer.writerow(row)
        
        file_size = file_path.stat().st_size
        
        return {
            'success': True,
            'file_path': str(file_path),
            'file_name': file_path.name,
            'file_size': file_size,
            'ads_count': len(ads),
            'platform': 'Google Ads',
            'format_note': 'Listo para importar en Google Ads Editor'
        }
    
    # =========================================================================
    # EXPORTACIÓN FACEBOOK ADS
    # =========================================================================
    
    def _export_facebook_ads(
        self,
        ads: List[Dict[str, Any]],
        filename: str,
        campaign_name: str = "Nueva Campaña",
        ad_set_name: str = "Conjunto de Anuncios 1",
        destination_url: str = "https://example.com",
        call_to_action: str = "LEARN_MORE",
        status: str = "PAUSED"
    ) -> Dict[str, Any]:
        """
        Exporta en formato compatible con Facebook Ads.
        
        Args:
            ads: Lista de anuncios
            campaign_name: Nombre de la campaña
            ad_set_name: Nombre del conjunto de anuncios
            destination_url: URL de destino
            call_to_action: CTA (LEARN_MORE, SHOP_NOW, etc.)
            status: Estado inicial
        """
        file_path = self.export_dir / f"{filename}_facebook_ads.csv"
        
        with open(file_path, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.DictWriter(f, fieldnames=self.FACEBOOK_ADS_COLUMNS)
            writer.writeheader()
            
            for i, ad in enumerate(ads, 1):
                # Facebook usa el primer headline como título
                # y la primera description como texto principal
                headlines = ad.get('headlines', [])
                descriptions = ad.get('descriptions', [])
                
                row = {
                    'Campaign Name': campaign_name,
                    'Ad Set Name': ad_set_name,
                    'Ad Name': f"Anuncio {i}",
                    'Primary Text': descriptions[0] if descriptions else '',
                    'Headline': headlines[0] if headlines else '',
                    'Description': descriptions[1] if len(descriptions) > 1 else '',
                    'Destination URL': destination_url,
                    'Call to Action': call_to_action,
                    'Status': status
                }
                
                writer.writerow(row)
        
        file_size = file_path.stat().st_size
        
        return {
            'success': True,
            'file_path': str(file_path),
            'file_name': file_path.name,
            'file_size': file_size,
            'ads_count': len(ads),
            'platform': 'Facebook Ads',
            'format_note': 'Compatible con Facebook Ads Manager'
        }
    
    # =========================================================================
    # EXPORTACIÓN ZIP
    # =========================================================================
    
    def _export_zip(
        self,
        ads: List[Dict[str, Any]],
        filename: str,
        include_formats: Optional[List[str]] = None
    ) -> Dict[str, Any]:
        """
        Exporta a múltiples formatos en un archivo ZIP.
        
        Args:
            ads: Lista de anuncios
            filename: Nombre base del archivo
            include_formats: Formatos a incluir (None = todos)
        """
        if not include_formats:
            include_formats = ['csv', 'json', 'txt', 'google_ads']
        
        zip_path = self.export_dir / f"{filename}.zip"
        
        # Crear ZIP
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            files_added = []
            
            for format in include_formats:
                if format not in self.SUPPORTED_FORMATS or format == 'zip':
                    continue
                
                try:
                    # Exportar a formato temporal
                    temp_filename = f"{filename}_{format}"
                    
                    if format == 'csv':
                        result = self._export_csv(ads, temp_filename)
                    elif format == 'json':
                        result = self._export_json(ads, temp_filename)
                    elif format == 'txt':
                        result = self._export_txt(ads, temp_filename)
                    elif format == 'google_ads':
                        result = self._export_google_ads(ads, temp_filename)
                    elif format == 'facebook_ads':
                        result = self._export_facebook_ads(ads, temp_filename)
                    else:
                        continue
                    
                    if result['success']:
                        # Agregar al ZIP
                        file_path = Path(result['file_path'])
                        zipf.write(file_path, file_path.name)
                        files_added.append(file_path.name)
                        
                        # Eliminar archivo temporal
                        file_path.unlink()
                
                except Exception as e:
                    logger.warning(f"⚠️ No se pudo incluir formato {format}: {e}")
        
        file_size = zip_path.stat().st_size
        
        return {
            'success': True,
            'file_path': str(zip_path),
            'file_name': zip_path.name,
            'file_size': file_size,
            'ads_count': len(ads),
            'formats_included': include_formats,
            'files_in_zip': files_added
        }
    
    # =========================================================================
    # EXPORTACIÓN BULK
    # =========================================================================
    
    def export_bulk(
        self,
        ads_groups: Dict[str, List[Dict[str, Any]]],
        format: str = 'csv'
    ) -> Dict[str, Any]:
        """
        Exporta múltiples grupos de anuncios.
        
        Args:
            ads_groups: Diccionario con nombre_grupo: lista_de_anuncios
            format: Formato de exportación
        
        Returns:
            Resultado de exportación bulk
        """
        logger.info(f"📦 Exportación bulk de {len(ads_groups)} grupos")
        
        results = []
        
        for group_name, ads in ads_groups.items():
            logger.info(f"   Exportando grupo: {group_name}")
            
            result = self.export(
                ads=ads,
                format=format,
                filename=group_name
            )
            
            results.append({
                'group_name': group_name,
                'result': result
            })
        
        # Resumen
        successful = sum(1 for r in results if r['result']['success'])
        failed = len(results) - successful
        
        bulk_result = {
            'success': failed == 0,
            'total_groups': len(ads_groups),
            'successful': successful,
            'failed': failed,
            'results': results,
            'exported_at': datetime.now().isoformat()
        }
        
        logger.info(f"✅ Exportación bulk completada: {successful}/{len(ads_groups)} exitosos")
        
        return bulk_result
    
    # =========================================================================
    # VALIDACIÓN
    # =========================================================================
    
    def _validate_ads(self, ads: List[Dict[str, Any]]) -> Dict[str, Any]:
        """
        Valida anuncios antes de exportar.
        
        Args:
            ads: Lista de anuncios
        
        Returns:
            Resultado de validación
        """
        issues = []
        warnings = []
        
        for i, ad in enumerate(ads, 1):
            # Validar estructura básica
            if 'headlines' not in ad and 'descriptions' not in ad:
                issues.append(f"Anuncio {i}: No tiene headlines ni descriptions")
            
            # Validar headlines
            if 'headlines' in ad:
                headlines = ad['headlines']
                
                if not headlines:
                    warnings.append(f"Anuncio {i}: Lista de headlines vacía")
                
                for j, headline in enumerate(headlines, 1):
                    if len(headline) > 30:
                        warnings.append(f"Anuncio {i}, Headline {j}: Excede 30 caracteres")
            
            # Validar descriptions
            if 'descriptions' in ad:
                descriptions = ad['descriptions']
                
                if not descriptions:
                    warnings.append(f"Anuncio {i}: Lista de descriptions vacía")
                
                for j, desc in enumerate(descriptions, 1):
                    if len(desc) > 90:
                        warnings.append(f"Anuncio {i}, Description {j}: Excede 90 caracteres")
        
        valid = len(issues) == 0
        
        return {
            'valid': valid,
            'issues': issues,
            'warnings': warnings,
            'warning': f"{len(issues)} problemas y {len(warnings)} advertencias encontradas" if issues or warnings else "Validación exitosa"
        }
    
    # =========================================================================
    # UTILIDADES
    # =========================================================================
    
    def _generate_filename(self, format: str) -> str:
        """Genera nombre de archivo automático."""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        return f"anuncios_{timestamp}"
    
    def _add_timestamp(self, filename: str) -> str:
        """Agrega timestamp a nombre de archivo."""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        return f"{filename}_{timestamp}"
    
    def _update_stats(self, result: Dict[str, Any]) -> None:
        """Actualiza estadísticas."""
        if result['success']:
            self.stats['total_exports'] += 1
            self.stats['total_files_created'] += 1
            
            format = result.get('format', 'unknown')
            self.stats['exports_by_format'][format] = \
                self.stats['exports_by_format'].get(format, 0) + 1
            
            file_size = result.get('file_size', 0)
            self.stats['total_bytes_exported'] += file_size
    
    def get_export_history(
        self,
        limit: Optional[int] = None
    ) -> List[Dict[str, Any]]:
        """Obtiene historial de exportaciones."""
        history = self.export_history.copy()
        
        if limit:
            history = history[-limit:]
        
        return history
    
    def get_statistics(self) -> Dict[str, Any]:
        """Obtiene estadísticas del gestor."""
        mb_exported = self.stats['total_bytes_exported'] / 1024 / 1024
        
        return {
            **self.stats,
            'mb_exported': round(mb_exported, 2),
            'avg_size_per_file': round(
                self.stats['total_bytes_exported'] / self.stats['total_files_created']
            ) if self.stats['total_files_created'] > 0 else 0
        }
    
    def list_exports(self) -> List[Dict[str, Any]]:
        """Lista archivos exportados en el directorio."""
        exports = []
        
        for file_path in self.export_dir.glob('*'):
            if file_path.is_file():
                exports.append({
                    'file_name': file_path.name,
                    'file_path': str(file_path),
                    'size': file_path.stat().st_size,
                    'created_at': datetime.fromtimestamp(
                        file_path.stat().st_ctime
                    ).isoformat()
                })
        
        # Ordenar por fecha de creación (más recientes primero)
        exports.sort(key=lambda x: x['created_at'], reverse=True)
        
        return exports
    
    def delete_export(self, filename: str) -> bool:
        """
        Elimina un archivo exportado.
        
        Args:
            filename: Nombre del archivo
        
        Returns:
            True si se eliminó exitosamente
        """
        file_path = self.export_dir / filename
        
        if file_path.exists() and file_path.is_file():
            try:
                file_path.unlink()
                logger.info(f"🗑️ Archivo eliminado: {filename}")
                return True
            except Exception as e:
                logger.error(f"❌ Error eliminando archivo: {e}")
                return False
        
        return False
    
    def clean_old_exports(self, days: int = 30) -> Dict[str, int]:
        """
        Limpia exportaciones antiguas.
        
        Args:
            days: Eliminar archivos más antiguos que X días
        
        Returns:
            Diccionario con resultados de limpieza
        """
        from datetime import timedelta
        
        cutoff_date = datetime.now() - timedelta(days=days)
        
        deleted = 0
        total_size = 0
        
        for file_path in self.export_dir.glob('*'):
            if file_path.is_file():
                file_time = datetime.fromtimestamp(file_path.stat().st_ctime)
                
                if file_time < cutoff_date:
                    try:
                        size = file_path.stat().st_size
                        file_path.unlink()
                        deleted += 1
                        total_size += size
                    except Exception as e:
                        logger.warning(f"⚠️ No se pudo eliminar {file_path.name}: {e}")
        
        logger.info(f"🗑️ Limpieza completada: {deleted} archivos eliminados")
        
        return {
            'deleted_files': deleted,
            'bytes_freed': total_size,
            'mb_freed': round(total_size / 1024 / 1024, 2)
        }


# ============================================================================
# FUNCIONES DE UTILIDAD
# ============================================================================

def create_export_manager(
    export_dir: Optional[str] = None,
    default_format: str = 'csv'
) -> ExportManager:
    """
    Factory function para crear gestor de exportación.
    
    Args:
        export_dir: Directorio de exportación
        default_format: Formato por defecto
    
    Returns:
        Instancia de ExportManager
    """
    return ExportManager(
        export_dir=export_dir,
        default_format=default_format
    )


# ============================================================================
# EJEMPLO DE USO
# ============================================================================

if __name__ == "__main__":
    print("="*60)
    print("📦 EXPORT MANAGER - Ejemplo de Uso")
    print("="*60)
    
    # Crear gestor
    manager = ExportManager()
    
    # Anuncios de ejemplo
    sample_ads = [
        {
            'id': 1,
            'headlines': [
                'Amarres de Amor Efectivos',
                'Recupera a Tu Pareja Ya',
                'Brujería Profesional'
            ],
            'descriptions': [
                'Amarres de amor con resultados garantizados en 24 horas.',
                'Bruja profesional con experiencia. Consulta gratis.'
            ],
            'tone': 'emocional',
            'keywords': ['amarres de amor', 'hechizos', 'brujería']
        },
        {
            'id': 2,
            'headlines': [
                'Tarot Profesional Online',
                'Lectura de Cartas Precisa',
                'Consulta tu Futuro'
            ],
            'descriptions': [
                'Tarot profesional con años de experiencia.',
                'Predicciones precisas. Primera consulta gratis.'
            ],
            'tone': 'profesional',
            'keywords': ['tarot', 'lectura de cartas', 'videncia']
        }
    ]
    
    print(f"\n📊 Anuncios de ejemplo: {len(sample_ads)}")
    
    # Exportar a CSV
    print(f"\n📄 Exportando a CSV...")
    result_csv = manager.export(sample_ads, format='csv', filename='anuncios_test')
    
    if result_csv['success']:
        print(f"   ✅ CSV creado: {result_csv['file_name']}")
        print(f"   📊 Filas: {result_csv['rows']}, Columnas: {result_csv['columns']}")
        print(f"   💾 Tamaño: {result_csv['file_size']} bytes")
    
    # Exportar a JSON
    print(f"\n📄 Exportando a JSON...")
    result_json = manager.export(sample_ads, format='json', filename='anuncios_test')
    
    if result_json['success']:
        print(f"   ✅ JSON creado: {result_json['file_name']}")
    
    # Exportar para Google Ads
    print(f"\n📄 Exportando para Google Ads...")
    result_google = manager.export(
        sample_ads,
        format='google_ads',
        filename='anuncios_test',
        campaign_name='Campaña Esotérico',
        ad_group_name='Amarres y Tarot',
        final_url='https://example.com'
    )
    
    if result_google['success']:
        print(f"   ✅ Google Ads CSV creado: {result_google['file_name']}")
        print(f"   📝 {result_google['format_note']}")
    
    # Exportar a ZIP (múltiples formatos)
    print(f"\n📄 Exportando a ZIP (múltiples formatos)...")
    result_zip = manager.export(
        sample_ads,
        format='zip',
        filename='anuncios_test',
        include_formats=['csv', 'json', 'txt', 'google_ads']
    )
    
    if result_zip['success']:
        print(f"   ✅ ZIP creado: {result_zip['file_name']}")
        print(f"   📦 Formatos incluidos: {', '.join(result_zip['formats_included'])}")
        print(f"   📁 Archivos: {', '.join(result_zip['files_in_zip'])}")
    
    # Listar exportaciones
    print(f"\n📂 ARCHIVOS EXPORTADOS:")
    exports = manager.list_exports()
    for i, exp in enumerate(exports[:5], 1):
        size_kb = exp['size'] / 1024
        print(f"   {i}. {exp['file_name']} ({size_kb:.1f} KB)")
    
    # Estadísticas
    print(f"\n📊 ESTADÍSTICAS:")
    stats = manager.get_statistics()
    print(f"   Total exportaciones: {stats['total_exports']}")
    print(f"   Archivos creados: {stats['total_files_created']}")
    print(f"   MB exportados: {stats['mb_exported']:.2f}")
    print(f"   Por formato: {stats['exports_by_format']}")
    
    print("\n" + "="*60)
    print("✅ Ejemplo completado")
    print("="*60)
    
    print(f"\n📁 Archivos guardados en: {manager.export_dir}")